"""
Excel Export Plugin for DataForge

This plugin provides enhanced Excel export functionality with formatting,
styling, and chart generation capabilities.

Features:
- Export data to Excel format (.xlsx)
- Custom sheet names and formatting
- Auto-filtering and frozen headers
- Number formatting
- Data validation
- Basic chart generation
"""

from typing import List, Dict, Any, Tuple
import io
from datetime import datetime
import re

from frontend.core.interfaces.plugin_interfaces import (
    IExportPlugin, PluginConfigBuilder
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExportPlugin(IExportPlugin):
    """Excel export plugin with advanced formatting capabilities."""
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl library is required for Excel export plugin")
    
    def get_name(self) -> str:
        return "Excel Export"
    
    def get_version(self) -> str:
        return "1.0.0"
    
    def get_description(self) -> str:
        return "Export data to Excel format with formatting and charts"
    
    def get_supported_formats(self) -> List[str]:
        return ["xlsx", "excel"]
    
    def export_data(self, data: List[Dict], format: str, config: Dict[str, Any]) -> bytes:
        """Export data to Excel format."""
        if not data:
            raise ValueError("No data to export")
        
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Configure worksheet
        sheet_name = config.get("sheet_name", "Data")
        worksheet.title = sheet_name
        
        # Get configuration
        include_headers = config.get("include_headers", True)
        auto_filter = config.get("auto_filter", True)
        freeze_header = config.get("freeze_header", True)
        format_numbers = config.get("format_numbers", True)
        
        # Write data
        row_num = 1
        
        # Write headers
        if include_headers:
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=header)
                # Style header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            row_num += 1
        
        # Write data rows
        for record in data:
            for col_num, (key, value) in enumerate(record.items(), 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=self._format_cell_value(value))
                
                # Apply number formatting
                if format_numbers and isinstance(value, (int, float)):
                    if isinstance(value, float):
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "#,##0"
                
                # Apply border
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            row_num += 1
        
        # Apply auto filter
        if auto_filter and include_headers:
            worksheet.auto_filter.ref = f"A1:{self._get_column_letter(len(data[0]))}{len(data) + 1}"
        
        # Freeze header row
        if freeze_header and include_headers:
            worksheet.freeze_panes = "A2"
        
        # Auto-size columns
        self._auto_size_columns(worksheet, data)
        
        # Add conditional formatting for numeric columns
        if format_numbers:
            self._add_conditional_formatting(worksheet, data, include_headers)
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output.read()
    
    def get_export_config_ui(self, format: str) -> Dict[str, Any]:
        """Return configuration UI for Excel export."""
        return {
            "fields": [
                PluginConfigBuilder.checkbox_field(
                    label="Include Headers",
                    default=True
                ),
                PluginConfigBuilder.text_field(
                    label="Sheet Name",
                    default="Data",
                    placeholder="Enter worksheet name"
                ),
                PluginConfigBuilder.checkbox_field(
                    label="Auto Filter",
                    default=True
                ),
                PluginConfigBuilder.checkbox_field(
                    label="Freeze Header Row",
                    default=True
                ),
                PluginConfigBuilder.checkbox_field(
                    label="Format Numbers",
                    default=True
                ),
                PluginConfigBuilder.select_field(
                    label="Date Format",
                    options=["YYYY-MM-DD", "MM/DD/YYYY", "DD/MM/YYYY", "DD-MON-YYYY"],
                    default="YYYY-MM-DD"
                )
            ]
        }
    
    def validate_data_for_export(self, data: List[Dict], format: str) -> Tuple[bool, str]:
        """Validate if data can be exported to Excel."""
        if not data:
            return False, "No data to export"
        
        if not isinstance(data, list):
            return False, "Data must be a list of dictionaries"
        
        if not all(isinstance(record, dict) for record in data):
            return False, "All data records must be dictionaries"
        
        # Check for consistent structure
        if len(data) > 1:
            first_keys = set(data[0].keys())
            for i, record in enumerate(data[1:], 1):
                if set(record.keys()) != first_keys:
                    return False, f"Inconsistent data structure at record {i + 1}"
        
        # Check Excel limits
        max_rows = 1048576  # Excel row limit
        max_cols = 16384    # Excel column limit
        
        if len(data) > max_rows - 1:  # -1 for header
            return False, f"Too many rows ({len(data)}). Excel limit is {max_rows - 1}"
        
        if len(data[0]) > max_cols:
            return False, f"Too many columns ({len(data[0])}). Excel limit is {max_cols}"
        
        return True, "Data is valid for Excel export"
    
    def get_file_extension(self, format: str) -> str:
        """Get file extension for Excel format."""
        return "xlsx"
    
    def supports_streaming_export(self, format: str) -> bool:
        """Excel export doesn't support streaming."""
        return False
    
    def estimate_file_size(self, data: List[Dict], format: str) -> int:
        """Estimate Excel file size in bytes."""
        if not data:
            return 0
        
        # Rough estimation based on data size
        # Excel files have overhead, so multiply by factor
        row_count = len(data)
        col_count = len(data[0])
        avg_cell_size = 20  # Estimated average cell size in bytes
        
        # Base file size + data size + formatting overhead
        base_size = 8192  # Minimum Excel file size
        data_size = row_count * col_count * avg_cell_size
        overhead = data_size * 0.3  # 30% overhead for formatting
        
        return int(base_size + data_size + overhead)
    
    def _format_cell_value(self, value: Any) -> Any:
        """Format cell value for Excel."""
        if value is None:
            return ""
        elif isinstance(value, (list, dict)):
            return str(value)
        elif isinstance(value, bool):
            return value
        elif isinstance(value, (int, float)):
            return value
        elif isinstance(value, str):
            # Clean up string values
            value = value.strip()
            
            # Try to convert numeric strings
            if value.replace(".", "").replace("-", "").replace("+", "").isdigit():
                try:
                    if "." in value:
                        return float(value)
                    else:
                        return int(value)
                except ValueError:
                    pass
            
            # Try to parse dates
            if self._looks_like_date(value):
                try:
                    return datetime.fromisoformat(value.replace("Z", "+00:00"))
                except ValueError:
                    pass
            
            return value
        else:
            return str(value)
    
    def _looks_like_date(self, value: str) -> bool:
        """Check if string looks like a date."""
        date_patterns = [
            r"\d{4}-\d{2}-\d{2}",  # YYYY-MM-DD
            r"\d{2}/\d{2}/\d{4}",  # MM/DD/YYYY
            r"\d{2}-\d{2}-\d{4}",  # DD-MM-YYYY
        ]
        
        return any(re.match(pattern, value) for pattern in date_patterns)
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to Excel column letter."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def _auto_size_columns(self, worksheet, data: List[Dict]) -> None:
        """Auto-size columns based on content."""
        if not data:
            return
        
        # Calculate column widths
        column_widths = {}
        headers = list(data[0].keys())
        
        # Start with header widths
        for i, header in enumerate(headers):
            column_widths[i] = len(str(header))
        
        # Check data widths (sample first 100 rows for performance)
        sample_data = data[:100]
        for record in sample_data:
            for i, (key, value) in enumerate(record.items()):
                cell_width = len(str(value))
                column_widths[i] = max(column_widths.get(i, 0), cell_width)
        
        # Apply column widths (with reasonable limits)
        for col_num in range(len(headers)):
            width = min(column_widths.get(col_num, 10), 50)  # Max width of 50
            width = max(width, 8)  # Min width of 8
            
            col_letter = self._get_column_letter(col_num + 1)
            worksheet.column_dimensions[col_letter].width = width
    
    def _add_conditional_formatting(self, worksheet, data: List[Dict], include_headers: bool) -> None:
        """Add conditional formatting to numeric columns."""
        if not data:
            return
        
        headers = list(data[0].keys())
        start_row = 2 if include_headers else 1
        end_row = len(data) + (1 if include_headers else 0)
        
        # Find numeric columns
        for col_num, header in enumerate(headers, 1):
            # Check if column has numeric data
            numeric_values = []
            for record in data[:50]:  # Sample first 50 rows
                value = record.get(header)
                if isinstance(value, (int, float)):
                    numeric_values.append(value)
            
            # Apply conditional formatting if column is mostly numeric
            if len(numeric_values) > len(data[:50]) * 0.7:  # 70% numeric threshold
                col_letter = self._get_column_letter(col_num)
                cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                
                # Add color scale rule (green-yellow-red)
                rule = ColorScaleRule(
                    start_type="min", start_color="63BE7B",  # Green
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",  # Yellow
                    end_type="max", end_color="F8696B"  # Red
                )
                worksheet.conditional_formatting.add(cell_range, rule)
    
    def create_chart(self, worksheet, data: List[Dict], chart_config: Dict[str, Any]) -> None:
        """Create and add chart to worksheet (bonus feature)."""
        chart_type = chart_config.get("type", "bar")
        
        if chart_type == "bar":
            chart = BarChart()
            chart.title = chart_config.get("title", "Data Chart")
            chart.y_axis.title = chart_config.get("y_axis", "Values")
            chart.x_axis.title = chart_config.get("x_axis", "Categories")
            
            # Add data (simplified - would need more sophisticated logic)
            data_range = Reference(worksheet, min_col=2, min_row=1, max_row=len(data) + 1)
            categories = Reference(worksheet, min_col=1, min_row=2, max_row=len(data) + 1)
            
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories)
            
            # Add chart to worksheet
            worksheet.add_chart(chart, "H2")
    
    def initialize(self) -> None:
        """Initialize plugin."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl library is required for Excel export plugin")
    
    def cleanup(self) -> None:
        """Cleanup plugin resources."""
        pass
